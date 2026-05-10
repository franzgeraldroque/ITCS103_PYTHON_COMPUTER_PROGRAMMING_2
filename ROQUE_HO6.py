import openpyxl as op

def favorite_people():
    workbook = op.Workbook()
    sheet = workbook.active

    sheet["A1"] = "ID"
    sheet["B1"] = "First Name"
    sheet["C1"] = "Last Name"
    sheet["D1"] = "Birth Year"
    sheet["E1"] = "Age"

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 13
    sheet.column_dimensions["C"].width = 13
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 10

    print("===== Favorite People Recorder =====")

    for i in range (1,4):
        print(f"\nFavorite Person {i}")
        fname = input(f"Enter the first name of your favorite person number {i}: ")
        lname = input(f"Enter the last name of your favorite person number {i}: ")
        birth_year = int(input(f"Enter the birth year of your favorite person number {i}: "))
        age = 2026 - birth_year

        person_data = [i, fname, lname, birth_year, age]
        sheet.append(person_data)

    workbook.save("favorite_people.xlsx")
    print("\nFavorite people added successfulley!")

    print("\n===== Favorite People List =====\n")
    workbook_read = op.load_workbook("favorite_people.xlsx")
    worksheet_read = workbook_read.active

    for row in worksheet_read.iter_rows(values_only=True):
        print(row)
    
    input("Please, press Enter to exit......")

favorite_people()