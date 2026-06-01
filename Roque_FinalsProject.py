import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op


window = tk.Tk()
window.title("Fitness Progress Tracker")
window.configure(bg="lightblue")

def display():
    workbook = op.load_workbook("Roque_Database.xlsx")
    sheet = workbook.active

    for row in table.get_children():
        table.delete(row)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        table.insert("", tk.END, values=row)

def input_validation():
    first = fname_entry.get()
    last = lname_entry.get()
    age = age_entry.get()
    height = height_entry.get()
    weight = weight_entry.get()
    streak = streak_entry.get()

    if not first or not last or not age or not height or not weight or not streak:
        messagebox.showerror("ERROR", "All field required!!!")
        return False
    
    if not height.isdigit() or not weight.isdigit() or not age.isdigit() or not streak.isdigit():
        messagebox.showerror("ERROR", "Age, Height, Weight, and Streak Days must be a number!!")
        return False
    
    return True

def streak_milestone():
    streak = int(streak_entry.get())

    if streak == 1:
        messagebox.showinfo("Milestone!", "Good job! Starting is better than doing nothing")

    elif streak == 7:
        messagebox.showinfo("Milestone!", "One week strong! Your foundation is built, keep the momentum going!!\n\nREMINDER: It has been a week! Please select your record and Update your current weight.")

    elif streak == 10:
        messagebox.showinfo("Milestone!", "Double digits achieved! Great work, keep it going!!\n\nREMINDER: It has been 10 days! It's time to Update your current weight again, to check your current BMI.")

    elif streak == 15:
        messagebox.showinfo("Milestone!", "Halfway mark reached! Your discipline is remarkable, keep pushing!!\n\nREMINDER: Halfway done! Update your current weight to keep in track of your BMI.")

    elif streak == 20:
        messagebox.showinfo("Milestone!", "Woww!! Most people quit here; stay resilient, you're too close to stop now!!\n\nREMINDER: Well disciplined! Keep in track by Updating your weight today.")

    elif streak == 30:
        messagebox.showinfo("Milestone!", "Milestone unlocked! 30 days of dedication and discipline, be proud and reward you're self!!\n\nREMINDER: What a Milestone! It's End of the Month, please Update your weight to see your overall progress.")

def bmi_chart(event=None):
    height_t = height_entry.get()
    weight_t = weight_entry.get()

    if height_t == "" or weight_t == "":
        return

    height = int(height_entry.get())
    weight = int(weight_entry.get())

    height_cal = height / 100
    bmi = round(weight / (height_cal * height_cal), 2)

    if bmi <= 18.5:
        status = "Underweight"

    elif bmi > 18.5 and bmi < 25:
        status = "Normal Weight"

    elif bmi >= 25:
        status = "Overweight or Obese"

    bmi_label.config(text=f"Your BMI is {bmi}, which means you're {status}!\n<18.5 = Underweight\n18.6 - 24.9 = Normal Weight\n>25 = Overweight")

def create(): 
    if not input_validation():
        return
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    age = int(age_entry.get())
    height = int(height_entry.get())
    weight = int(weight_entry.get())
    streak = int(streak_entry.get())

    height_cal = height / 100
    bmi = round(weight / (height_cal * height_cal), 2)

    workbook = op.load_workbook("Roque_Database.xlsx")
    sheet = workbook.active

    create_id = sheet.max_row
    
    sheet.append([create_id, last, first, middle, age, height, weight, bmi, streak])
    workbook.save("Roque_Database.xlsx")

    messagebox.showinfo("Success", "Record added successfully")
    
    bmi_chart()

    streak_milestone()

    display()

def auto_populate(event):
    selected = table.focus()
    values = table.item(selected, "values")

    if values:
        fname_entry.delete(0, tk.END)
        mname_entry.delete(0, tk.END)
        lname_entry.delete(0, tk.END)
        age_entry.delete(0, tk.END)
        height_entry.delete(0, tk.END)
        weight_entry.delete(0, tk.END)
        streak_entry.delete(0, tk.END)

        fname_entry.insert(0, values[2])
        mname_entry.insert(0, values[3])
        lname_entry.insert(0, values[1])
        age_entry.insert(0, values[4])
        height_entry.insert(0, values[5])
        weight_entry.insert(0, values[6])
        streak_entry.insert(0, values[8])

def update():
    selected = table.focus()

    if not selected:
        messagebox.showerror("ERROR", "Select a record first!!!")

    if not input_validation():
        return
    
    values = table.item(selected, "values")
    record_id = values[0]

    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    age = int(age_entry.get())
    height = int(height_entry.get())
    weight = int(weight_entry.get())
    streak = int(streak_entry.get())

    height_cal = height / 100
    bmi = round(weight / (height_cal * height_cal), 2)

    workbook = op.load_workbook("Roque_Database.xlsx")
    sheet = workbook.active

    for rows in sheet.iter_rows(min_row=2):
        if str(rows[0].value) == str(record_id):
            rows[1].value = last
            rows[3].value = middle
            rows[2].value = first
            rows[4].value = age
            rows[5].value = height
            rows[6].value = weight
            rows[7].value = bmi
            rows[8].value = streak
    
    workbook.save("Roque_Database.xlsx")
    messagebox.showinfo("Success", "Record updated Successfully")

    bmi_chart()

    streak_milestone()

    display()

def delete():
    selected = table.focus()

    if not selected:
        messagebox.showerror("ERROR", "Select a record first!!!")
    
    values = table.item(selected, "values")
    record_id = values[0]

    confirm = messagebox.askyesnocancel("Confirm", "Are you sure you want to delete this record?")
    if not confirm:
        return
    
    workbook = op.load_workbook("Roque_Database.xlsx")
    sheet = workbook.active

    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == str(record_id):
            sheet.delete_rows(i)
            break

    workbook.save("Roque_Database.xlsx")
    messagebox.showinfo("Success", "Record deleted successfully")
    
    display()

# Form Title
title = tk.Label(window, text="Fitness Progress Tracker", font=("Times New Roman", 14, "bold"), bg="lightblue")
title.grid(row=0, column=0, columnspan=9)

# Frame
genframe = tk.Frame(window, bg="lightblue", bd=4, relief="groove")
genframe.grid(row=1, column=0, columnspan=9, padx=10, pady=10)

# First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins", 12))
fname_entry.grid(row=2, column=1, columnspan=1, padx=10, pady=(10, 0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins", 10, "italic"), bg="lightblue")
fname_label.grid(row=3, column=1, columnspan=1)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins", 12))
mname_entry.grid(row=2, column=2, columnspan=1, padx=10, pady=(10, 0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins", 10, "italic"), bg="lightblue")
mname_label.grid(row=3, column=2, columnspan=1)

# Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins", 12))
lname_entry.grid(row=2, column=3, columnspan=1, padx=10, pady=(10, 0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins", 10, "italic"), bg="lightblue")
lname_label.grid(row=3, column=3, columnspan=1)

# Age Entry
age_entry = tk.Entry(genframe, font=("Poppins", 12))
age_entry.grid(row=4, column=1, columnspan=1, padx=10, pady=(10, 0))

age_label = tk.Label(genframe, text="Age", font=("Poppins", 10, "italic"), bg="lightblue")
age_label.grid(row=5, column=1, columnspan=1)

# Height Entry (In cm)
height_entry = tk.Entry(genframe, font=("Poppins", 12))
height_entry.grid(row=4, column=2, columnspan=1, padx=10, pady=(10, 0))

height_label = tk.Label(genframe, text="Height (In cm)", font=("Poppins", 10, "italic"), bg="lightblue")
height_label.grid(row=5, column=2, columnspan=1)

# Weight Entry (In kg)
weight_entry = tk.Entry(genframe, font=("Poppins", 12))
weight_entry.grid(row=4, column=3, columnspan=1, padx=10, pady=(10, 0))

weight_label = tk.Label(genframe, text="Weight (In kg)", font=("Poppins", 10, "italic"), bg="lightblue")
weight_label.grid(row=5, column=3, columnspan=1)

# Streak Days
streak_entry = tk.Entry(genframe, font=("Poppins", 12))
streak_entry.grid(row=6, column=0, columnspan=2, padx=10, pady=(10, 0))

streak_label = tk.Label(genframe, text="Streak Days", font=("Poppins", 10, "italic"), bg="lightblue")
streak_label.grid(row=7, column=0, columnspan=2)

bmi_label = tk.Label(genframe,  text="BMI:\n<18.5 = Underweight\n18.6 - 24.9 = Normal Weight\n>25 = Overweight", font=("Poppins", 12, "italic"), bg="lightblue")
bmi_label.grid(row=6, column=2, columnspan=2, rowspan=2)
bmi_label.bind("<Return>", bmi_chart)

# Buttons
submit_btn = tk.Button(window, text="Submit", font=("Poppins", 12, "bold"), bg="lightpink", command=create)
submit_btn.grid(row=6, column=1, columnspan=5, pady=(10, 20))

update_btn = tk.Button(window, text="Update",font=("Poppins", 12, "bold"), bg="lightgreen", command=update)
update_btn.grid(row=6, column=2, columnspan=5)

delete_btn = tk.Button(window, text="Delete", bg="red", fg="white",font=("Poppins", 12, "bold"), command=delete)
delete_btn.grid(row=6, column=3, columnspan=5)

# Table
table = ttk.Treeview(
    window,
    columns=("Customer ID", "Last Name","First Name", "Middle Name", "Age", "Height", "Weight", "BMI", "Streak Days"),
    show="headings"
)

for headings in ("Customer ID", "Last Name","First Name", "Middle Name", "Age", "Height", "Weight", "BMI", "Streak Days"):
    table.heading(headings, text=headings)

table.grid(row=7, column=0, columnspan=9, padx=10, pady=10)

table.bind("<<TreeviewSelect>>", auto_populate)

display()
window.mainloop()
